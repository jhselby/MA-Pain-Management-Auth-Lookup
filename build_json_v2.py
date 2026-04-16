#!/usr/bin/env python3
"""
Build JSON files for the MA Pain Management Authorization Lookup site.

Expected layout:
- This script lives in the same folder as the workbook
- The workbook is named: data.xlsx

Usage:
    python build_json_v2.py
    python build_json_v2.py /path/to/data.xlsx

Outputs:
- rules.json
- evidence_map.json
- sources.json
- metadata.json

This version adds:
- payer-level public metrics/profile data (where available)
- computed friction_score / friction_label / friction_reasons
- no required UI changes
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_SHEETS = {
    "Normalized_Data",
    "Evidence_Map",
    "Source_Register",
}

PAYER_METRICS: dict[str, dict[str, Any]] = {
    "MassHealth": {
        "metrics_source": "MassHealth Calendar Year 2025 Prior Authorization Metrics",
        "metrics_as_of": "2026-04",
        "avg_standard_turnaround_days": 4.21,
        "avg_expedited_turnaround_days": 1.33,
        "denial_signal": "Moderate",
        "strictness_signal": "Moderate",
        "notes": (
            "Public metrics now available. Use as payer-level friction context only, "
            "not CPT-specific denial intelligence."
        ),
    },
    "United": {
        "metrics_source": "UnitedHealthcare Prior Authorization Utilization Review Statistics",
        "metrics_as_of": "2026-04",
        "avg_standard_turnaround_days": None,
        "avg_expedited_turnaround_days": None,
        "denial_signal": "Moderate",
        "strictness_signal": "Moderate-High",
        "notes": (
            "Public utilization review statistics exist, but they are broad and not "
            "Massachusetts pain-procedure specific."
        ),
    },
}


def slugify(value: str) -> str:
    value = (value or "").strip().lower()
    out = []
    for ch in value:
        if ch.isalnum():
            out.append(ch)
        else:
            out.append("-")
    slug = "".join(out)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug.strip("-")


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def ws_to_dicts(ws) -> list[dict[str, Any]]:
    header = [normalize_value(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [normalize_value(cell) for cell in row]
        if not any(v not in ("", None) for v in values):
            continue
        rows.append(dict(zip(header, values)))
    return rows


def split_semicolon_list(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in str(value).split(";") if item and item.strip()]


def compute_friction(
    *,
    auth_status: str,
    auth_vendor: str,
    confidence: str,
    booking_note: str,
    evidence_count: int,
    source_count: int,
    payer: str,
    procedure_group: str,
) -> tuple[int, str, list[str]]:
    score = 0
    reasons: list[str] = []

    auth_status_l = (auth_status or "").strip().lower()
    auth_vendor_l = (auth_vendor or "").strip().lower()
    confidence_l = (confidence or "").strip().lower()
    booking_note_l = (booking_note or "").strip().lower()

    if auth_status_l == "yes":
        score += 2
        reasons.append("Auth required")
    elif auth_status_l == "verify":
        score += 1
        reasons.append("Coverage needs verification")

    vendor_terms = [
        "evicore", "carelon", "aim", "evolent", "interqual",
        "authorization manager", "portal", "code checker", "api", "edi 278"
    ]
    if any(term in auth_vendor_l for term in vendor_terms):
        score += 1
        reasons.append("Vendor or portal workflow")

    if confidence_l == "medium":
        score += 1
        reasons.append("Moderate certainty")
    elif confidence_l in {"low", "verify"}:
        score += 2
        reasons.append("Lower certainty")

    ambiguity_terms = [
        "verify", "not clearly", "depends", "exact cpt", "exact setting",
        "exact product", "site-of-service", "site of service", "product complexity"
    ]
    if any(term in booking_note_l for term in ambiguity_terms):
        score += 1
        reasons.append("Ambiguous public guidance")

    if evidence_count == 0:
        score += 1
        reasons.append("No linked evidence row")
    if source_count == 0:
        score += 1
        reasons.append("No linked source register entry")

    procedure_l = (procedure_group or "").lower()
    if "stimulator" in procedure_l or "pump" in procedure_l or "vertebro" in procedure_l or "kypho" in procedure_l:
        score += 1
        reasons.append("Higher-complexity procedure family")

    metrics = PAYER_METRICS.get(payer, {})
    denial_signal = str(metrics.get("denial_signal", "")).lower()
    strictness_signal = str(metrics.get("strictness_signal", "")).lower()
    avg_std_days = metrics.get("avg_standard_turnaround_days")

    if "high" in denial_signal:
        score += 2
        reasons.append("Higher denial environment")
    elif "moderate" in denial_signal:
        score += 1
        reasons.append("Moderate denial environment")

    if "high" in strictness_signal:
        score += 2
        reasons.append("Higher strictness payer")
    elif "moderate" in strictness_signal:
        score += 1
        reasons.append("Moderate strictness payer")

    if isinstance(avg_std_days, (int, float)):
        if avg_std_days > 5:
            score += 2
            reasons.append("Slower standard turnaround")
        elif avg_std_days > 3:
            score += 1
            reasons.append("Moderate standard turnaround")

    deduped: list[str] = []
    for reason in reasons:
        if reason not in deduped:
            deduped.append(reason)

    if score <= 1:
        label = "Low"
    elif score <= 3:
        label = "Moderate"
    elif score <= 5:
        label = "High"
    else:
        label = "Very High"

    return score, label, deduped


def build_outputs(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=True)

    missing = REQUIRED_SHEETS.difference(wb.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {sorted(missing)}")

    normalized = ws_to_dicts(wb["Normalized_Data"])
    evidence_rows = ws_to_dicts(wb["Evidence_Map"])
    source_rows = ws_to_dicts(wb["Source_Register"])

    sources = []
    evidence_map = []
    evidence_by_id: dict[str, dict[str, Any]] = {}
    evidence_index: dict[tuple[str, str], list[str]] = {}

    for row in source_rows:
        source = {
            "source_id": row.get("source_id", ""),
            "payer_group": row.get("payer_group", ""),
            "source_type": row.get("source_type", ""),
            "url": row.get("url", ""),
            "what_it_supports": row.get("what_it_supports", ""),
            "source_owner": row.get("source_owner", ""),
            "source_level": row.get("source_level", ""),
            "expected_update_cycle": row.get("expected_update_cycle", ""),
            "access_method": row.get("access_method", ""),
            "recommended_verification_use": row.get("recommended_verification_use", ""),
            "added_by": row.get("added_by", ""),
            "added_on": row.get("added_on", ""),
            "notes": row.get("notes", ""),
        }
        sources.append(source)

    for row in evidence_rows:
        source_ids = split_semicolon_list(row.get("source_bundle", ""))
        evidence = {
            "evidence_id": row.get("evidence_id", ""),
            "payer": row.get("payer", ""),
            "procedure_group": row.get("procedure_group", ""),
            "procedure_slug": slugify(row.get("procedure_group", "")),
            "cpt_family": row.get("cpt_family", ""),
            "field_validated": row.get("field_validated", ""),
            "value_in_matrix": row.get("value_in_matrix", ""),
            "confidence": row.get("confidence", ""),
            "source_ids": source_ids,
            "verification_method": row.get("verification_method", ""),
            "rationale_summary": row.get("rationale_summary", ""),
            "public_limitations": row.get("public_limitations", ""),
            "recommended_reviewer_action": row.get("recommended_reviewer_action", ""),
            "source_urls": split_semicolon_list(row.get("source_urls", "")),
            "last_verified": row.get("last_verified", ""),
            "notes": row.get("notes", ""),
        }
        evidence_map.append(evidence)
        if evidence["evidence_id"]:
            evidence_by_id[evidence["evidence_id"]] = evidence
        key = (evidence["payer"], evidence["procedure_group"])
        evidence_index.setdefault(key, []).append(evidence["evidence_id"])

    procedures: dict[str, Any] = {}
    payers = set()

    for row in normalized:
        payer = row.get("payer", "")
        procedure_group = row.get("procedure_group", "")
        if not payer or not procedure_group:
            continue

        payers.add(payer)
        proc_slug = slugify(procedure_group)
        procedures.setdefault(proc_slug, {
            "procedure_group": procedure_group,
            "procedure_slug": proc_slug,
            "cpt_family": row.get("cpt_family", ""),
            "payers": {},
        })

        linked_evidence_ids = evidence_index.get((payer, procedure_group), [])
        linked_source_ids: list[str] = []
        for evid in linked_evidence_ids:
            match = evidence_by_id.get(evid)
            if match:
                for sid in match["source_ids"]:
                    if sid not in linked_source_ids:
                        linked_source_ids.append(sid)

        friction_score, friction_label, friction_reasons = compute_friction(
            auth_status=row.get("auth_status", ""),
            auth_vendor=row.get("auth_vendor", ""),
            confidence=row.get("confidence", ""),
            booking_note=row.get("booking_note", ""),
            evidence_count=len(linked_evidence_ids),
            source_count=len(linked_source_ids),
            payer=payer,
            procedure_group=procedure_group,
        )

        procedures[proc_slug]["payers"][payer] = {
            "auth": row.get("auth_status", ""),
            "sedation_type": row.get("sedation_default", ""),
            "auth_vendor": row.get("auth_vendor", ""),
            "do_not_book": row.get("booking_note", ""),
            "confidence": row.get("confidence", ""),
            "last_verified": row.get("last_verified", ""),
            "state_scope": row.get("state_scope", ""),
            "implementation_note": row.get("implementation_note", ""),
            "evidence_ids": linked_evidence_ids,
            "source_ids": linked_source_ids,
            "payer_metrics": PAYER_METRICS.get(payer, {}),
            "friction_score": friction_score,
            "friction_label": friction_label,
            "friction_reasons": friction_reasons,
        }

    now_utc = datetime.now(timezone.utc).isoformat()

    rules = {
        "dataset_name": "MA Pain Management Authorization Matrix",
        "generated_at": now_utc,
        "generated_from": workbook_path.name,
        "payers": sorted(payers),
        "procedures": procedures,
    }

    metadata = {
        "dataset_name": "MA Pain Management Authorization Matrix",
        "generated_at": now_utc,
        "generated_from": workbook_path.name,
        "workbook_expected_name_for_local_edits": "data.xlsx",
        "coverage_scope": "Massachusetts pain management payers",
        "sheets_used": sorted(REQUIRED_SHEETS),
        "payer_count": len(payers),
        "procedure_count": len(procedures),
        "metrics_layer_notes": (
            "Friction fields are operational burden signals derived from auth requirement, "
            "vendor workflow, ambiguity, evidence/source linkage, and limited public payer-level metrics."
        ),
        "public_metrics_payers": sorted(PAYER_METRICS.keys()),
        "warning": (
            "Use as decision support. Always verify member/product/site-of-service "
            "details in the payer or delegated-vendor portal before booking."
        ),
        "local_testing_tip": (
            "If index.html is opened directly and JSON fails to load, run "
            "'python -m http.server' in this folder and open http://localhost:8000."
        ),
    }

    return {
        "rules": rules,
        "evidence_map": evidence_map,
        "sources": sources,
        "metadata": metadata,
    }


def write_json(path: Path, payload: Any) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    default_workbook = base_dir / "data.xlsx"

    if len(sys.argv) > 1:
        workbook_path = Path(sys.argv[1]).expanduser().resolve()
    else:
        workbook_path = default_workbook.resolve()

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        print("Tip: either rename your workbook to data.xlsx or pass the workbook path explicitly.")
        return 1

    try:
        outputs = build_outputs(workbook_path)
    except Exception as exc:
        print(f"Build failed: {exc}")
        return 2

    write_json(base_dir / "rules.json", outputs["rules"])
    write_json(base_dir / "evidence_map.json", outputs["evidence_map"])
    write_json(base_dir / "sources.json", outputs["sources"])
    write_json(base_dir / "metadata.json", outputs["metadata"])

    print("Build complete.")
    print(f"Workbook: {workbook_path.name}")
    print("Generated:")
    for name in ("rules.json", "evidence_map.json", "sources.json", "metadata.json"):
        print(f" - {base_dir / name}")
    print(f"Payers: {outputs['metadata']['payer_count']}")
    print(f"Procedures: {outputs['metadata']['procedure_count']}")
    print(f"Public metrics payers: {', '.join(outputs['metadata']['public_metrics_payers']) or 'None'}")
    print(f"Generated at (UTC): {outputs['metadata']['generated_at']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
