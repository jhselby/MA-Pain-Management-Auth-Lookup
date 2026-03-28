#!/usr/bin/env python3
"""
Build JSON files for the MA Pain Management Authorization Lookup site.

Expected layout:
- This script lives in the same folder as the workbook
- The workbook is named: data.xlsx

Usage:
    python build_json.py
    python build_json.py /path/to/data.xlsx

Outputs:
- rules.json
- evidence_map.json
- sources.json
- metadata.json
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_SHEETS = {
    "Normalized_Data",
    "Evidence_Map",
    "Source_Register",
}


def slugify(value: str) -> str:
    value = (value or "").strip().lower()
    out = []
    for ch in value:
        if ch.isalnum():
            out.append(ch)
        else:
            out.append("-")
    slug = "".join(out)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug.strip("-")


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def ws_to_dicts(ws) -> list[dict[str, Any]]:
    header = [normalize_value(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [normalize_value(cell) for cell in row]
        if not any(v not in ("", None) for v in values):
            continue
        rows.append(dict(zip(header, values)))
    return rows


def split_semicolon_list(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in str(value).split(";") if item and item.strip()]


def build_outputs(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=True)

    missing = REQUIRED_SHEETS.difference(wb.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {sorted(missing)}")

    normalized = ws_to_dicts(wb["Normalized_Data"])
    evidence_rows = ws_to_dicts(wb["Evidence_Map"])
    source_rows = ws_to_dicts(wb["Source_Register"])

    # Build sources lookup.
    sources = []
    sources_by_id = {}
    for row in source_rows:
        source = {
            "source_id": row.get("source_id", ""),
            "payer_group": row.get("payer_group", ""),
            "source_type": row.get("source_type", ""),
            "url": row.get("url", ""),
            "what_it_supports": row.get("what_it_supports", ""),
            "source_owner": row.get("source_owner", ""),
            "source_level": row.get("source_level", ""),
            "expected_update_cycle": row.get("expected_update_cycle", ""),
            "access_method": row.get("access_method", ""),
            "recommended_verification_use": row.get("recommended_verification_use", ""),
            "added_by": row.get("added_by", ""),
            "added_on": row.get("added_on", ""),
            "notes": row.get("notes", ""),
        }
        sources.append(source)
        if source["source_id"]:
            sources_by_id[source["source_id"]] = source

    # Build evidence rows.
    evidence_map = []
    evidence_index: dict[tuple[str, str], list[str]] = {}
    for row in evidence_rows:
        source_ids = split_semicolon_list(row.get("source_bundle", ""))
        evidence = {
            "evidence_id": row.get("evidence_id", ""),
            "payer": row.get("payer", ""),
            "procedure_group": row.get("procedure_group", ""),
            "procedure_slug": slugify(row.get("procedure_group", "")),
            "cpt_family": row.get("cpt_family", ""),
            "field_validated": row.get("field_validated", ""),
            "value_in_matrix": row.get("value_in_matrix", ""),
            "confidence": row.get("confidence", ""),
            "source_ids": source_ids,
            "verification_method": row.get("verification_method", ""),
            "rationale_summary": row.get("rationale_summary", ""),
            "public_limitations": row.get("public_limitations", ""),
            "recommended_reviewer_action": row.get("recommended_reviewer_action", ""),
            "source_urls": split_semicolon_list(row.get("source_urls", "")),
            "last_verified": row.get("last_verified", ""),
            "notes": row.get("notes", ""),
        }
        evidence_map.append(evidence)
        key = (evidence["payer"], evidence["procedure_group"])
        evidence_index.setdefault(key, []).append(evidence["evidence_id"])

    # Build nested rules object for the website.
    procedures: dict[str, Any] = {}
    payers = set()

    for row in normalized:
        payer = row.get("payer", "")
        procedure_group = row.get("procedure_group", "")
        if not payer or not procedure_group:
            continue

        payers.add(payer)
        proc_slug = slugify(procedure_group)
        procedures.setdefault(proc_slug, {
            "procedure_group": procedure_group,
            "procedure_slug": proc_slug,
            "cpt_family": row.get("cpt_family", ""),
            "payers": {},
        })

        linked_evidence_ids = evidence_index.get((payer, procedure_group), [])
        linked_source_ids = []
        for evid in linked_evidence_ids:
            match = next((e for e in evidence_map if e["evidence_id"] == evid), None)
            if match:
                for sid in match["source_ids"]:
                    if sid not in linked_source_ids:
                        linked_source_ids.append(sid)

        procedures[proc_slug]["payers"][payer] = {
            "auth": row.get("auth_status", ""),
            "sedation_type": row.get("sedation_default", ""),
            "auth_vendor": row.get("auth_vendor", ""),
            "do_not_book": row.get("booking_note", ""),
            "confidence": row.get("confidence", ""),
            "last_verified": row.get("last_verified", ""),
            "state_scope": row.get("state_scope", ""),
            "implementation_note": row.get("implementation_note", ""),
            "evidence_ids": linked_evidence_ids,
            "source_ids": linked_source_ids,
        }

    rules = {
        "dataset_name": "MA Pain Management Authorization Matrix",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generated_from": workbook_path.name,
        "payers": sorted(payers),
        "procedures": procedures,
    }

    metadata = {
        "dataset_name": "MA Pain Management Authorization Matrix",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generated_from": workbook_path.name,
        "workbook_expected_name_for_local_edits": "data.xlsx",
        "coverage_scope": "Massachusetts pain management payers",
        "sheets_used": sorted(REQUIRED_SHEETS),
        "payer_count": len(payers),
        "procedure_count": len(procedures),
        "warning": (
            "Use as decision support. Always verify member/product/site-of-service "
            "details in the payer or delegated-vendor portal before booking."
        ),
        "local_testing_tip": (
            "If index.html is opened directly and JSON fails to load, run "
            "'python -m http.server' in this folder and open http://localhost:8000."
        ),
    }

    return {
        "rules": rules,
        "evidence_map": evidence_map,
        "sources": sources,
        "metadata": metadata,
    }


def write_json(path: Path, payload: Any) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    default_workbook = base_dir / "data.xlsx"

    if len(sys.argv) > 1:
        workbook_path = Path(sys.argv[1]).expanduser().resolve()
    else:
        workbook_path = default_workbook.resolve()

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        print("Tip: either rename your workbook to data.xlsx or pass the workbook path explicitly.")
        return 1

    try:
        outputs = build_outputs(workbook_path)
    except Exception as exc:
        print(f"Build failed: {exc}")
        return 2

    write_json(base_dir / "rules.json", outputs["rules"])
    write_json(base_dir / "evidence_map.json", outputs["evidence_map"])
    write_json(base_dir / "sources.json", outputs["sources"])
    write_json(base_dir / "metadata.json", outputs["metadata"])

    print("Build complete.")
    print(f"Workbook: {workbook_path.name}")
    print("Generated:")
    for name in ("rules.json", "evidence_map.json", "sources.json", "metadata.json"):
        print(f" - {base_dir / name}")
    print(f"Payers: {outputs['metadata']['payer_count']}")
    print(f"Procedures: {outputs['metadata']['procedure_count']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
